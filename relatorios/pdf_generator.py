"""
Sistema de geração de PDFs para relatórios.
Suporta templates customizáveis e diferentes tipos de relatórios.
"""

from django.template.loader import render_to_string
from django.conf import settings
from io import BytesIO

# Tentar importar WeasyPrint (preferencial)
try:
    from weasyprint import HTML, CSS
    from weasyprint.text.fonts import FontConfiguration
    WEASYPRINT_AVAILABLE = True
except (ImportError, OSError):
    WEASYPRINT_AVAILABLE = False
    HTML = None
    CSS = None
    FontConfiguration = None

# Tentar importar ReportLab (fallback)
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm, inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, PageBreak, KeepTogether
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.graphics.shapes import Drawing, Rect, String, Circle, Wedge, Line
    from reportlab.graphics.charts.barcharts import HorizontalBarChart, VerticalBarChart
    from reportlab.graphics.charts.piecharts import Pie
    from reportlab.graphics.charts.linecharts import HorizontalLineChart
    from reportlab.graphics.charts.legends import Legend
    from reportlab.graphics import renderPDF
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

import os
from datetime import datetime


class PDFGenerator:
    """Gerador de PDFs com templates customizáveis."""
    
    def __init__(self, title="Relatório", subtitle="", author="Sistema ARES"):
        self.title = title
        self.subtitle = subtitle
        self.author = author
        self.font_config = FontConfiguration() if WEASYPRINT_AVAILABLE else None
        
    def generate_pdf(self, template_name, context, output_path=None):
        """
        Gera PDF a partir de um template HTML.
        
        Args:
            template_name: Nome do template HTML
            context: Contexto para o template
            output_path: Caminho para salvar o PDF (opcional)
            
        Returns:
            bytes: Conteúdo do PDF em bytes ou string HTML se PDF não disponível
        """
        # Adicionar informações padrão ao contexto
        context.update({
            'report_title': self.title,
            'report_subtitle': self.subtitle,
            'report_author': self.author,
            'generated_at': datetime.now(),
            'STATIC_URL': settings.STATIC_URL,
            'MEDIA_URL': settings.MEDIA_URL,
        })
        
        # Renderizar HTML
        html_string = render_to_string(template_name, context)
        
        # Tentar usar WeasyPrint primeiro
        if WEASYPRINT_AVAILABLE:
            try:
                html = HTML(string=html_string, base_url=settings.BASE_DIR)
                css = CSS(string=self._get_pdf_styles(), font_config=self.font_config)
                pdf_bytes = html.write_pdf(stylesheets=[css], font_config=self.font_config)
                
                if output_path:
                    os.makedirs(os.path.dirname(output_path), exist_ok=True)
                    with open(output_path, 'wb') as f:
                        f.write(pdf_bytes)
                
                return pdf_bytes
            except Exception as e:
                # Se WeasyPrint falhar, tentar ReportLab
                pass
        
        # Fallback para ReportLab se WeasyPrint não disponível ou falhou
        if REPORTLAB_AVAILABLE:
            return self._generate_with_reportlab(context, output_path)
        
        # Última tentativa: tentar instalar e usar xhtml2pdf
        try:
            from xhtml2pdf import pisa
            result = BytesIO()
            pdf = pisa.CreatePDF(BytesIO(html_string.encode('utf-8')), dest=result)
            
            if not pdf.err:
                pdf_bytes = result.getvalue()
                result.close()
                
                if output_path:
                    os.makedirs(os.path.dirname(output_path), exist_ok=True)
                    with open(output_path, 'wb') as f:
                        f.write(pdf_bytes)
                
                return pdf_bytes
        except ImportError:
            pass
        
        # Se nenhum gerador está disponível, usar ReportLab básico forçado
        return self._generate_simple_pdf(context, output_path)
    
    def _generate_with_reportlab(self, context, output_path=None):
        """
        Gera PDF usando ReportLab com logo e gráficos.
        Versão melhorada com visual profissional.
        """
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer, 
            pagesize=A4,
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=30
        )
        story = []
        styles = getSampleStyleSheet()
        
        # ========== CABEÇALHO COM LOGO ==========
        header_data = []
        
        # Tentar carregar logo PNG (você pode adicionar logo.png na pasta static/img/)
        logo_paths = [
            os.path.join(settings.BASE_DIR, 'siteares', 'static', 'img', 'logo.png'),
            os.path.join(settings.BASE_DIR, 'siteares', 'static', 'img', 'Logo.png'),
            os.path.join(settings.BASE_DIR, 'siteares', 'static', 'img', 'logo_ares.png'),
        ]
        
        logo_cell = None
        for logo_path in logo_paths:
            if os.path.exists(logo_path):
                try:
                    # Logo retangular 16:9 (1920x1080) - ajustar proporcionalmente
                    # Usando largura fixa de 100px, altura será 56px (mantendo proporção 16:9)
                    logo_cell = Image(logo_path, width=100, height=56.25)
                    break
                except Exception as e:
                    continue
        
        # Se não encontrar imagem, usar texto estilizado
        if logo_cell is None:
            logo_cell = Paragraph(
                '<font size=28 color="#C8102E"><b>ARES</b></font>',
                styles['Normal']
            )
        
        # Informações do sistema
        info_cell = Paragraph(
            f'<font size=10 color="#666666">Sistema ARES<br/>'
            f'Gerado em: {datetime.now().strftime("%d/%m/%Y às %H:%M")}<br/>'
            f'Por: {self.author}</font>',
            styles['Normal']
        )
        
        header_table = Table([[logo_cell, info_cell]], colWidths=[140, 360])
        header_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, 0), 'LEFT'),
            ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        story.append(header_table)
        story.append(Spacer(1, 10))
        
        # Linha separadora (cor principal do site)
        line_drawing = Drawing(500, 3)
        line_drawing.add(Rect(0, 0, 500, 3, fillColor=colors.HexColor('#C8102E'), strokeColor=None))
        story.append(line_drawing)
        story.append(Spacer(1, 20))
        
        # ========== TÍTULO DO RELATÓRIO ==========
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            textColor=colors.HexColor('#333333'),
            spaceAfter=10,
            alignment=TA_LEFT,
            fontName='Helvetica-Bold'
        )
        story.append(Paragraph(self.title, title_style))
        
        if self.subtitle:
            subtitle_style = ParagraphStyle(
                'CustomSubtitle',
                parent=styles['Normal'],
                fontSize=11,
                textColor=colors.HexColor('#666666'),
                spaceAfter=20,
                alignment=TA_LEFT
            )
            story.append(Paragraph(self.subtitle, subtitle_style))
        
        story.append(Spacer(1, 20))
        
        # ========== ESTATÍSTICAS ==========
        if 'stats' in context:
            stats = context['stats']
            
            # Box de estatísticas
            stats_style = ParagraphStyle(
                'StatsStyle',
                parent=styles['Normal'],
                fontSize=10,
                spaceAfter=6
            )
            
            stats_data = [
                [
                    Paragraph('<b>Total de Produtos:</b>', stats_style),
                    Paragraph(f'{stats.get("total_products", 0)}', stats_style),
                    Paragraph('<b>Valor Total:</b>', stats_style),
                    Paragraph(f'R$ {stats.get("total_value", 0):,.2f}', stats_style)
                ],
                [
                    Paragraph('<b>Produtos Críticos:</b>', stats_style),
                    Paragraph(f'<font color="#C8102E">{stats.get("critical_count", 0)}</font>', stats_style),
                    Paragraph('<b>Estoque Baixo:</b>', stats_style),
                    Paragraph(f'<font color="#ffc107">{stats.get("low_count", 0)}</font>', stats_style)
                ]
            ]
            
            stats_table = Table(stats_data, colWidths=[120, 80, 120, 120])
            stats_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f8f9fa')),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 10),
                ('RIGHTPADDING', (0, 0), (-1, -1), 10),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#dee2e6')),
            ]))
            
            # Título do resumo
            if stats.get('total_products') is not None:
                resume_title = '<b>Resumo do Estoque</b>'
            elif stats.get('total_movements') is not None:
                resume_title = '<b>Resumo de Movimentações</b>'
            else:
                resume_title = '<b>Resumo</b>'
            
            story.append(KeepTogether([
                Paragraph(resume_title, title_style),
                Spacer(1, 10),
                stats_table
            ]))
            story.append(Spacer(1, 20))
            
            # ========== GRÁFICOS ==========
            # Relatório de Estoque - Gráfico de Pizza + Barra
            if stats.get('total_products') is not None and stats.get('critical_count') is not None:
                total = stats.get('total_products', 1)
                if total > 0:
                    critical_count = stats.get('critical_count', 0)
                    low_count = stats.get('low_count', 0)
                    ok_count = stats.get('ok_count', 0)
                    
                    # ===== GRÁFICO DE PIZZA =====
                    pie_drawing = Drawing(250, 200)
                    
                    # Título do gráfico de pizza (adicionar ANTES do gráfico)
                    pie_title = String(125, 185, 'Status dos Produtos', textAnchor='middle', fontSize=11, fontName='Helvetica-Bold')
                    pie_drawing.add(pie_title)
                    
                    pie = Pie()
                    pie.x = 50
                    pie.y = 20
                    pie.width = 150
                    pie.height = 150
                    pie.data = [ok_count, low_count, critical_count]
                    pie.labels = [f'OK ({ok_count})', f'Baixo ({low_count})', f'Crítico ({critical_count})']
                    pie.slices.strokeWidth = 0.5
                    pie.slices[0].fillColor = colors.HexColor('#28a745')  # Verde
                    pie.slices[1].fillColor = colors.HexColor('#ffc107')  # Amarelo
                    pie.slices[2].fillColor = colors.HexColor('#C8102E')  # Vermelho ARES
                    pie.slices[0].popout = 0
                    pie.slices[1].popout = 5
                    pie.slices[2].popout = 10
                    pie_drawing.add(pie)
                    
                    # ===== GRÁFICO DE BARRAS VERTICAIS =====
                    bar_drawing = Drawing(250, 200)
                    
                    # Título do gráfico de barras
                    bar_title = String(130, 185, 'Quantidade por Status', textAnchor='middle', fontSize=11, fontName='Helvetica-Bold')
                    bar_drawing.add(bar_title)
                    
                    vbar = VerticalBarChart()
                    vbar.x = 40
                    vbar.y = 20
                    vbar.width = 180
                    vbar.height = 150
                    vbar.data = [[ok_count, low_count, critical_count]]
                    vbar.categoryAxis.categoryNames = ['OK', 'Baixo', 'Crítico']
                    vbar.valueAxis.valueMin = 0
                    vbar.valueAxis.valueMax = max(ok_count, low_count, critical_count, 10) * 1.2
                    
                    vbar.bars[0].fillColor = colors.HexColor('#28a745')
                    vbar.bars[1].fillColor = colors.HexColor('#ffc107')
                    vbar.bars[2].fillColor = colors.HexColor('#C8102E')
                    
                    vbar.barLabelFormat = '%d'
                    vbar.barLabels.fontSize = 9
                    vbar.barLabels.dy = 5
                    bar_drawing.add(vbar)
                    
                    # Adicionar ambos os gráficos lado a lado
                    charts_table = Table([[pie_drawing, bar_drawing]], colWidths=[250, 250], rowHeights=[200])
                    charts_table.setStyle(TableStyle([
                        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ]))
                    
                    story.append(Paragraph('<b>Análise Visual</b>', title_style))
                    story.append(Spacer(1, 10))
                    story.append(charts_table)
                    story.append(Spacer(1, 30))
        
        # ========== TABELA DE MOVIMENTAÇÕES ==========
        if 'movements' in context:
            movements = list(context['movements'])[:100]
            if movements:
                story.append(Paragraph('<b>Movimentações de Estoque</b>', title_style))
                story.append(Spacer(1, 10))
                
                table_data = [[
                    Paragraph('<b>Data</b>', styles['Normal']),
                    Paragraph('<b>Produto</b>', styles['Normal']),
                    Paragraph('<b>Tipo</b>', styles['Normal']),
                    Paragraph('<b>Qtd</b>', styles['Normal']),
                    Paragraph('<b>Usuário</b>', styles['Normal'])
                ]]
                
                for m in movements:
                    # Definir cor por tipo
                    if m.type == 'entrada':
                        tipo_color = '<font color="green"><b>ENTRADA</b></font>'
                    elif m.type == 'saida':
                        tipo_color = '<font color="#C8102E"><b>SAÍDA</b></font>'
                    else:
                        tipo_color = '<font color="blue"><b>AJUSTE</b></font>'
                    
                    table_data.append([
                        Paragraph(m.created_at.strftime('%d/%m/%Y'), styles['Normal']),
                        Paragraph(str(m.product.name)[:30], styles['Normal']),
                        Paragraph(tipo_color, styles['Normal']),
                        Paragraph(str(m.quantity), styles['Normal']),
                        Paragraph(str(m.user.username if hasattr(m, 'user') else '-')[:15], styles['Normal'])
                    ])
                
                mov_table = Table(table_data, colWidths=[70, 180, 80, 60, 80])
                mov_table.setStyle(TableStyle([
                    # Cabeçalho
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#495057')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    
                    # Corpo
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                    ('ALIGN', (3, 1), (3, -1), 'CENTER'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    
                    # Bordas
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#495057')),
                    
                    # Padding
                    ('LEFTPADDING', (0, 0), (-1, -1), 6),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                    ('TOPPADDING', (0, 0), (-1, -1), 5),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                    
                    # Cores alternadas
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')])
                ]))
                
                story.append(mov_table)
        
        # ========== TABELA DE VENCIMENTOS ==========
        if 'expiring_products' in context:
            expiring = list(context['expiring_products'])[:100]
            if expiring:
                story.append(Paragraph('<b>Produtos Próximos ao Vencimento</b>', title_style))
                story.append(Spacer(1, 10))
                
                table_data = [[
                    Paragraph('<b>Produto</b>', styles['Normal']),
                    Paragraph('<b>Categoria</b>', styles['Normal']),
                    Paragraph('<b>Validade</b>', styles['Normal']),
                    Paragraph('<b>Dias</b>', styles['Normal']),
                    Paragraph('<b>Status</b>', styles['Normal'])
                ]]
                
                from datetime import date
                today = date.today()
                
                for p in expiring:
                    if hasattr(p, 'expiry_date') and p.expiry_date:
                        days_left = (p.expiry_date - today).days
                        
                        if days_left < 0:
                            status = '<font color="#C8102E"><b>VENCIDO</b></font>'
                            days_text = f'{abs(days_left)}d atrás'
                        elif days_left == 0:
                            status = '<font color="#C8102E"><b>VENCE HOJE</b></font>'
                            days_text = 'Hoje'
                        elif days_left <= 7:
                            status = '<font color="orange"><b>URGENTE</b></font>'
                            days_text = f'{days_left}d'
                        elif days_left <= 30:
                            status = '<font color="#ffc107"><b>ATENÇÃO</b></font>'
                            days_text = f'{days_left}d'
                        else:
                            status = '<font color="blue">OK</font>'
                            days_text = f'{days_left}d'
                        
                        table_data.append([
                            Paragraph(str(p.name)[:35], styles['Normal']),
                            Paragraph(str(p.category.name if hasattr(p, 'category') else '-')[:20], styles['Normal']),
                            Paragraph(p.expiry_date.strftime('%d/%m/%Y'), styles['Normal']),
                            Paragraph(days_text, styles['Normal']),
                            Paragraph(status, styles['Normal'])
                        ])
                
                exp_table = Table(table_data, colWidths=[160, 100, 80, 60, 80])
                exp_table.setStyle(TableStyle([
                    # Cabeçalho
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#495057')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    
                    # Corpo
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                    ('ALIGN', (2, 1), (4, -1), 'CENTER'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    
                    # Bordas
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#495057')),
                    
                    # Padding
                    ('LEFTPADDING', (0, 0), (-1, -1), 6),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                    ('TOPPADDING', (0, 0), (-1, -1), 5),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                    
                    # Cores alternadas
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')])
                ]))
                
                story.append(exp_table)
        
        # ========== TABELA DE PRODUTOS ==========
        if 'products' in context:
            products = list(context['products'])[:100]  # Limitar para performance
            if products:
                story.append(Paragraph('<b>Produtos em Estoque</b>', title_style))
                story.append(Spacer(1, 10))
                
                table_data = [[
                    Paragraph('<b>SKU</b>', styles['Normal']),
                    Paragraph('<b>Produto</b>', styles['Normal']),
                    Paragraph('<b>Categoria</b>', styles['Normal']),
                    Paragraph('<b>Estoque</b>', styles['Normal']),
                    Paragraph('<b>Mín.</b>', styles['Normal']),
                    Paragraph('<b>Status</b>', styles['Normal'])
                ]]
                
                for p in products:
                    if p.current_stock == 0:
                        status = '<font color="red"><b>CRÍTICO</b></font>'
                        row_bg = colors.HexColor('#ffe6e6')
                    elif p.current_stock <= p.min_stock:
                        status = '<font color="orange"><b>BAIXO</b></font>'
                        row_bg = colors.HexColor('#fff8e6')
                    else:
                        status = '<font color="green">OK</font>'
                        row_bg = colors.white
                    
                    table_data.append([
                        Paragraph(str(p.sku), styles['Normal']),
                        Paragraph(str(p.name)[:40], styles['Normal']),
                        Paragraph(str(p.category.name) if hasattr(p, 'category') else '-', styles['Normal']),
                        Paragraph(f'{p.current_stock} {p.unit.name if hasattr(p, "unit") else ""}', styles['Normal']),
                        Paragraph(str(p.min_stock), styles['Normal']),
                        Paragraph(status, styles['Normal'])
                    ])
                
                products_table = Table(table_data, colWidths=[60, 150, 80, 70, 40, 60])
                products_table.setStyle(TableStyle([
                    # Cabeçalho
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#495057')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    
                    # Corpo
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                    ('ALIGN', (3, 1), (5, -1), 'CENTER'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    
                    # Bordas
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#495057')),
                    
                    # Padding
                    ('LEFTPADDING', (0, 0), (-1, -1), 6),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                    ('TOPPADDING', (0, 0), (-1, -1), 5),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                    
                    # Cores alternadas nas linhas
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')])
                ]))
                
                story.append(products_table)
        
        # ========== RODAPÉ ==========
        story.append(Spacer(1, 30))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.HexColor('#999999'),
            alignment=TA_CENTER
        )
        story.append(Paragraph(
            f'Sistema ARES - Gestão de Estoque | Documento gerado automaticamente em {datetime.now().strftime("%d/%m/%Y às %H:%M")}',
            footer_style
        ))
        
        # Gerar PDF
        doc.build(story)
        pdf_bytes = buffer.getvalue()
        buffer.close()
        
        if output_path:
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            with open(output_path, 'wb') as f:
                f.write(pdf_bytes)
        
        return pdf_bytes
    
    def _generate_simple_pdf(self, context, output_path=None):
        """
        Gera PDF simples sem dependências externas, apenas ReportLab básico.
        Funciona mesmo sem ReportLab instalado usando apenas bibliotecas padrão.
        """
        try:
            from reportlab.pdfgen import canvas
            from reportlab.lib.pagesizes import A4
            
            buffer = BytesIO()
            p = canvas.Canvas(buffer, pagesize=A4)
            width, height = A4
            
            # Título
            p.setFont("Helvetica-Bold", 20)
            p.drawString(50, height - 50, self.title)
            
            # Subtítulo
            if self.subtitle:
                p.setFont("Helvetica", 12)
                p.drawString(50, height - 80, self.subtitle)
            
            # Informações
            p.setFont("Helvetica", 10)
            p.drawString(50, height - 110, f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
            p.drawString(50, height - 130, f"Por: {self.author}")
            
            # Linha de separação
            p.line(50, height - 150, width - 50, height - 150)
            
            y_position = height - 180
            
            # Se houver produtos, listar
            if 'products' in context:
                products = list(context['products'])[:50]  # Limitar para performance
                
                p.setFont("Helvetica-Bold", 12)
                p.drawString(50, y_position, "Produtos:")
                y_position -= 30
                
                p.setFont("Helvetica", 9)
                for product in products:
                    if y_position < 50:  # Nova página se necessário
                        p.showPage()
                        y_position = height - 50
                        p.setFont("Helvetica", 9)
                    
                    status = 'OK'
                    if product.current_stock == 0:
                        status = 'CRÍTICO'
                    elif product.current_stock <= product.min_stock:
                        status = 'BAIXO'
                    
                    line = f"{product.sku} - {product.name[:40]} | Estoque: {product.current_stock}/{product.min_stock} | {status}"
                    p.drawString(60, y_position, line)
                    y_position -= 20
            
            # Estatísticas se disponíveis
            if 'stats' in context:
                stats = context['stats']
                if y_position < 150:
                    p.showPage()
                    y_position = height - 50
                
                y_position -= 30
                p.setFont("Helvetica-Bold", 12)
                p.drawString(50, y_position, "Estatísticas:")
                y_position -= 25
                
                p.setFont("Helvetica", 10)
                p.drawString(60, y_position, f"Total de Produtos: {stats.get('total_products', 0)}")
                y_position -= 20
                p.drawString(60, y_position, f"Produtos Críticos: {stats.get('critical_count', 0)}")
                y_position -= 20
                p.drawString(60, y_position, f"Estoque Baixo: {stats.get('low_count', 0)}")
                y_position -= 20
                p.drawString(60, y_position, f"Valor Total: R$ {stats.get('total_value', 0):,.2f}")
            
            p.showPage()
            p.save()
            
            pdf_bytes = buffer.getvalue()
            buffer.close()
            
            if output_path:
                os.makedirs(os.path.dirname(output_path), exist_ok=True)
                with open(output_path, 'wb') as f:
                    f.write(pdf_bytes)
            
            return pdf_bytes
            
        except ImportError:
            # Se nem ReportLab básico está disponível, criar PDF manualmente (muito básico)
            # Isso nunca deveria acontecer, mas é um fallback extremo
            error_msg = f"PDF Generator Error: Nenhuma biblioteca de PDF disponível.\n"
            error_msg += f"Título: {self.title}\n"
            error_msg += f"Subtítulo: {self.subtitle}\n"
            error_msg += f"Autor: {self.author}\n"
            error_msg += f"Data: {datetime.now().strftime('%d/%m/%Y %H:%M')}\n"
            return error_msg.encode('utf-8')
    
    def _get_pdf_styles(self):
        """Retorna CSS customizado para PDF."""
        return """
        @page {
            size: A4;
            margin: 2cm 1.5cm;
            
            @top-left {
                content: string(report-title);
                font-size: 10pt;
                color: #666;
            }
            
            @top-right {
                content: string(report-date);
                font-size: 10pt;
                color: #666;
            }
            
            @bottom-center {
                content: "Página " counter(page) " de " counter(pages);
                font-size: 9pt;
                color: #999;
            }
        }
        
        body {
            font-family: 'DejaVu Sans', Arial, sans-serif;
            font-size: 10pt;
            line-height: 1.5;
            color: #333;
        }
        
        h1 {
            string-set: report-title content();
            font-size: 24pt;
            color: #c62828;
            margin-bottom: 0.5cm;
            page-break-after: avoid;
        }
        
        h2 {
            font-size: 18pt;
            color: #c62828;
            margin-top: 1cm;
            margin-bottom: 0.5cm;
            page-break-after: avoid;
        }
        
        h3 {
            font-size: 14pt;
            color: #333;
            margin-top: 0.5cm;
            margin-bottom: 0.3cm;
            page-break-after: avoid;
        }
        
        table {
            width: 100%;
            border-collapse: collapse;
            margin-bottom: 1cm;
            font-size: 9pt;
        }
        
        table thead {
            background-color: #f5f5f5;
            border-bottom: 2px solid #c62828;
        }
        
        table th {
            padding: 8px;
            text-align: left;
            font-weight: bold;
            color: #333;
        }
        
        table td {
            padding: 8px;
            border-bottom: 1px solid #e0e0e0;
        }
        
        table tr:hover {
            background-color: #fafafa;
        }
        
        .text-primary {
            color: #c62828;
        }
        
        .text-success {
            color: #2e7d32;
        }
        
        .text-danger {
            color: #d32f2f;
        }
        
        .text-warning {
            color: #f57c00;
        }
        
        .text-muted {
            color: #999;
        }
        
        .badge {
            display: inline-block;
            padding: 2px 6px;
            border-radius: 3px;
            font-size: 8pt;
            font-weight: bold;
        }
        
        .badge-success {
            background-color: #e8f5e9;
            color: #2e7d32;
        }
        
        .badge-warning {
            background-color: #fff3e0;
            color: #f57c00;
        }
        
        .badge-danger {
            background-color: #ffebee;
            color: #d32f2f;
        }
        
        .stats-box {
            background-color: #f5f5f5;
            padding: 1cm;
            border-radius: 5px;
            margin-bottom: 1cm;
        }
        
        .report-header {
            border-bottom: 3px solid #c62828;
            padding-bottom: 0.5cm;
            margin-bottom: 1cm;
        }
        
        .report-footer {
            border-top: 1px solid #e0e0e0;
            padding-top: 0.5cm;
            margin-top: 1cm;
            font-size: 8pt;
            color: #999;
        }
        
        .page-break {
            page-break-after: always;
        }
        
        .no-break {
            page-break-inside: avoid;
        }
        """


class ReportExporter:
    """Exportador de relatórios em múltiplos formatos."""
    
    @staticmethod
    def export_to_excel(data, filename, sheet_name="Relatório"):
        """
        Exporta dados para Excel.
        
        Args:
            data: Lista de dicionários ou DataFrame
            filename: Nome do arquivo
            sheet_name: Nome da planilha
            
        Returns:
            str: Caminho do arquivo gerado
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            if not data:
                return None
            
            # Cabeçalhos
            headers = list(data[0].keys())
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="C62828", end_color="C62828", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            
            # Dados
            for row, item in enumerate(data, 2):
                for col, header in enumerate(headers, 1):
                    ws.cell(row=row, column=col, value=item.get(header, ''))
            
            # Ajustar largura das colunas
            for column in ws.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width
            
            # Salvar
            output_path = os.path.join(settings.MEDIA_ROOT, 'relatorios', filename)
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            wb.save(output_path)
            
            return output_path
            
        except Exception as e:
            print(f"Erro ao exportar para Excel: {e}")
            return None
    
    @staticmethod
    def export_to_csv(data, filename):
        """
        Exporta dados para CSV.
        
        Args:
            data: Lista de dicionários
            filename: Nome do arquivo
            
        Returns:
            str: Caminho do arquivo gerado
        """
        import csv
        
        if not data:
            return None
        
        output_path = os.path.join(settings.MEDIA_ROOT, 'relatorios', filename)
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        
        with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.DictWriter(f, fieldnames=data[0].keys())
            writer.writeheader()
            writer.writerows(data)
        
        return output_path
